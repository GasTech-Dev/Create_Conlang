import openpyxl
from random import choice

def Creator():
    name_ofFile = "Trieur/Verbe_Franxois"#Marquer Le nom de la langue
    name_ofFile = name_ofFile + ".xlsx"
    wb_conjugue = openpyxl.Workbook()
    wb_langue = openpyxl.load_workbook(name_ofFile)
    feuille_conjugue = wb_conjugue['Sheet']
    feuille_verbe = wb_langue['Feuil1']
    voyelle = ["a", "e", "i", "o", "u", "y"]
    consomne = ["b", "c", "d", "f", "g", "h", "j", "k", "l", "m", "n", "p", "q", "r", "s", "t", "v", "w", "x", "z"]
    
    for row in range(1, feuille_verbe.max_row + 1):
        mot = feuille_verbe.cell(row=row, column=1).value

        termot = mot[-2:]
        
        if termot == "io":
            for i in range(3):
                verbe = mot[:-1]
                
                terminaison = choice(voyelle)
            
                verbe = verbe + terminaison
                print(verbe)
                feuille_conjugue.cell(row=row, column=1).value = verbe
            for i in range(3):
                verbe = mot[:-2]
                terminaison = choice(voyelle)
                terminaison = terminaison + choice(consomne)
                verbe = verbe + terminaison
                print(verbe)
            
                feuille_conjugue.cell(row=row, column=1).value = verbe
            print("\n\n")
    wb_conjugue.save("Trieur\Verbe_Conjuguer.xlsx")

def Groupe1():
    name_ofFile = "Trieur/Verbe_Franxois"#Marquer Le nom de la langue
    name_ofFile = name_ofFile + ".xlsx"
    wb_conjugue = openpyxl.Workbook()
    wb_langue = openpyxl.load_workbook(name_ofFile)
    feuille_conjugue = wb_conjugue['Sheet']
    feuille_verbe = wb_langue['Feuil1']
    conpteur = 1
    
    for row in range(1, feuille_verbe.max_row + 1):
        mot = feuille_verbe.cell(row=row, column=1).value

        termot = mot[-2:]
        
        if termot == "re":
            for i in range(3):
                verbe = mot[:-1]
                if i == 0:
                    terminaison = "a"
                elif i == 1:
                    terminaison = "y"
                else:
                    terminaison = "u"
            
                verbe = verbe + terminaison
                print(verbe)
                feuille_conjugue.cell(row=conpteur, column=1).value = verbe
                conpteur += 1
            for i in range(3):
                verbe = mot[:-2]
                if i == 0:
                    terminaison = "am"
                elif i == 1:
                    terminaison = "av"
                else:
                    terminaison = "az"
                verbe = verbe + terminaison
                feuille_conjugue.cell(row=conpteur, column=1).value = verbe
                conpteur += 1
            conpteur += 3
            
    wb_conjugue.save("Trieur\Verbe_Conjuguer.xlsx")
Groupe1()

