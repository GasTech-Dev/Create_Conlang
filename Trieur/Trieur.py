import openpyxl
from flask import Flask, request, render_template
import random

class Verbe():
    def trieur():
        name_ofFile = "Trieur/Franx"#Marquer Le nom de la langue

        name_ofFile = name_ofFile + ".xlsx"
        chemin_fichier =name_ofFile
        wb_langue = openpyxl.load_workbook(chemin_fichier)
        wb_verbe = openpyxl.load_workbook("Trieur/Verbe_Franxois.xlsx")
        feuille_langue = wb_langue['Sheet']
        feuille_verbe = wb_verbe['Feuil1']
        
        mottr = ""
        i = 1
        
            
        for row in range(1, feuille_langue.max_row + 1):
            MotLangue = feuille_langue.cell(row=row, column=1).value
            MotTraduit = feuille_langue.cell(row=row, column=2).value
            if MotLangue != None:             
                terminaison = MotLangue[-2:]
                if terminaison == "ir" or terminaison == "er":
                    
                    feuille_verbe.cell(row=i, column=2).value = MotLangue
                    feuille_verbe.cell(row=i, column=1).value = MotTraduit
                    print(MotLangue)
                    i += 1
                        

        wb_verbe.save("Trieur/Verbe_Franxois.xlsx")
        wb_langue.close()
        wb_verbe.close()
    def Analyseur():
        name_ofFile = "Trieur/Verbe_Franxois"#Marquer Le nom de la langue
        name_ofFile = name_ofFile + ".xlsx"
        wb_langue = openpyxl.load_workbook(name_ofFile)
        feuille_verbe = wb_langue['Feuil1']
        termot_l = []
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            termot = mot[-2:]
            termot_l.append(termot)

        #print(termot_l)
        first = []
        for i in termot_l:
            if i in first:
                pass
            else:
                first.append(i)
        for i in first:
            f = termot_l.count(i)
            
            if f > 80:
                print(i)
    def Rectifieur():
        name_ofFile = "Trieur/Verbe_Franxois"#Marquer Le nom de la langue
        name_ofFile = name_ofFile + ".xlsx"
        wb_langue = openpyxl.load_workbook(name_ofFile, data_only=True)
        feuille_verbe = wb_langue['Feuil1']
        termot_l = []
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            termot = mot[-2:]
            termot_l.append(termot)

        first = []
        for i in termot_l:
            if i in first:
                pass
            else:
                first.append(i)
        terminaison = []
        for i in first:
            f = termot_l.count(i)
            
            if f > 80:
                print(i)
                terminaison.append(i)
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            if mot is not None:
                termot = mot[-2:]
                if termot in terminaison:
                    pass
                else:
                    
                    mot = mot [:-1] + random.choice(terminaison)
                    print(mot)
                    feuille_verbe.cell(row=row, column=1).value = mot
                    #[:-1]
        wb_langue.save("Trieur/verbe_rectifier.xlsx")
        wb_langue.close()
class Nom():
    def trieur():
        name_ofFile = "Trieur/Nom_Franxois"#Marquer Le nom de la langue

        name_ofFile = name_ofFile + ".xlsx"
        chemin_fichier =name_ofFile
        wb_langue = openpyxl.load_workbook(chemin_fichier)
        wb_nom = openpyxl.load_workbook("Trieur/Nom_Franxois.xlsx")
        feuille_langue = wb_langue['Sheet']
        feuille_nom = wb_nom['Feuil1']
        
        mottr = ""
        i = 1
        
            
        for row in range(1, feuille_langue.max_row + 1):
            MotLangue = feuille_langue.cell(row=row, column=1).value
            MotTraduit = feuille_langue.cell(row=row, column=2).value
            if MotLangue != None:             
                terminaison = MotLangue[-2:]
                if terminaison == "ir" or terminaison=="er":
                    pass
                else:
                    feuille_nom.cell(row=i, column=2).value = MotLangue
                    feuille_nom.cell(row=i, column=1).value = MotTraduit
                    print(MotLangue)
                    i += 1


        wb_nom.save("Trieur/Nom_Franxois.xlsx")
        wb_langue.close()
        wb_nom.close()

    #Annalyse tout les nom et regarde nous donne qu'elle est la terminaison en commun dans chaques nom
    def Analyseur():
        name_ofFile = "Trieur/Nom_Franxois.xlsx"  # Fusionne les opérations pour éviter une erreur de chemin
        wb_langue = openpyxl.load_workbook(name_ofFile)
        feuille_verbe = wb_langue['Feuil1']
        termot_l = []
        
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            
            
            if mot is not None:
                termot = mot[-2:] 
                termot_l.append(termot)
            else:
            
                termot_l.append('')  



        first = []
        for i in termot_l:
            if i in first:
                pass
            else:
                first.append(i)
        for j in first:
            f = termot_l.count(j)
            
            if f > 1200:
                print(j)
                

        wb_langue.save("Trieur/Nom_Franxois.xlsx")
        wb_langue.close()
    def Rectifieur():
        name_ofFile = "Trieur/Nom_Franxois.xlsx"  # Fusionne les opérations pour éviter une erreur de chemin
        wb_langue = openpyxl.load_workbook(name_ofFile)
        feuille_verbe = wb_langue['Feuil1']
        termot_l = []
        
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            
            
            if mot is not None:
                termot = mot[-2:] 
                termot_l.append(termot)
            else:
            
                termot_l.append('')  



        first = []
        for i in termot_l:
            if i in first:
                pass
            else:
                first.append(i)
        terminaison = []
        for j in first:
            f = termot_l.count(j)
            
            if f > 1200:
                print(j)
                terminaison.append(j)
        for row in range(1, feuille_verbe.max_row + 1):
            mot = feuille_verbe.cell(row=row, column=1).value
            if mot is not None:
                if mot[-2:] in terminaison:
                    pass
                else:
                    mot = mot[:-2] + random.choice(terminaison)
                    feuille_verbe.cell(row=row, column=1).value = mot

        wb_langue.save("Trieur/Nom_Rectifier.xlsx")
        wb_langue.close()
Nom.Rectifieur()