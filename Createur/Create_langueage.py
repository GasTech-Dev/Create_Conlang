import openpyxl

from deep_translator import GoogleTranslator
import os
import random
# Ouvrir le fichier Excel

#crée une liste avec les consomne
consonants = ['b', 'c', 'd', 'f', 'g', 'h', 'j', 'k', 'l', 'm', 'n', 'p', 'q', 'r', 's', 't', 'v', 'w', 'x', 'y', 'z']
vowels = ['a', 'e', 'i', 'o', 'u']

def translate():
    wb = openpyxl.load_workbook('français.xlsx')
    rowe = 1
    # Sélectionner la feuille de calcul à utiliser
    for i in range(2):
        sheet = wb['Feuille1']
        print("afrikaans': 'af', 'albanian': 'sq', 'amharic': 'am', 'arabic': 'ar', 'armenian': 'hy', 'assamese': 'as', 'aymara': 'ay', 'azerbaijani': 'az', 'bambara': 'bm', 'basque': 'eu', 'belarusian': 'be', 'bengali': 'bn', 'bhojpuri': 'bho', 'bosnian': 'bs', 'bulgarian': 'bg', 'catalan': 'ca', 'cebuano': 'ceb', 'chichewa': 'ny', 'chinese (simplified)': 'zh-CN', 'chinese (traditional)': 'zh-TW', 'corsican': 'co', 'croatian': 'hr', 'czech': 'cs', 'danish': 'da', 'dhivehi': 'dv', 'dogri': 'doi', 'dutch': 'nl', 'english': 'en', 'esperanto': 'eo', 'estonian': 'et', 'ewe': 'ee', 'filipino': 'tl', 'finnish': 'fi', 'french': 'fr', 'frisian': 'fy', 'galician': 'gl', 'georgian': 'ka', 'german': 'de', 'greek': 'el', 'guarani': 'gn', 'gujarati': 'gu', 'haitian creole': 'ht', 'hausa': 'ha', 'hawaiian': 'haw', 'hebrew': 'iw', 'hindi': 'hi', 'hmong': 'hmn', 'hungarian': 'hu', 'icelandic': 'is', 'igbo': 'ig', 'ilocano': 'ilo', 'indonesian': 'id', 'irish': 'ga', 'italian': 'it', 'japanese': 'ja', 'javanese': 'jw', 'kannada': 'kn', 'kazakh': 'kk', 'khmer': 'km', 'kinyarwanda': 'rw', 'konkani': 'gom', 'korean': 'ko', 'krio': 'kri', 'kurdish (kurmanji)': 'ku', 'kurdish (sorani)': 'ckb', 'kyrgyz': 'ky', 'lao': 'lo', 'latin': 'la', 'latvian': 'lv', 'lingala': 'ln', 'lithuanian': 'lt', 'luganda': 'lg', 'luxembourgish': 'lb', 'macedonian': 'mk', 'maithili': 'mai', 'malagasy': 'mg', 'malay': 'ms', 'malayalam': 'ml', 'maltese': 'mt', 'maori': 'mi', 'marathi': 'mr', 'meiteilon (manipuri)': 'mni-Mtei', 'mizo': 'lus', 'mongolian': 'mn', 'myanmar': 'my', 'nepali': 'ne', 'norwegian': 'no', 'odia (oriya)': 'or', 'oromo': 'om', 'pashto': 'ps', 'persian': 'fa', 'polish': 'pl', 'portuguese': 'pt', 'punjabi': 'pa', 'quechua': 'qu', 'romanian': 'ro', 'russian': 'ru', 'samoan': 'sm', 'sanskrit': 'sa', 'scots gaelic': 'gd', 'sepedi': 'nso', 'serbian': 'sr', 'sesotho': 'st', 'shona': 'sn', 'sindhi': 'sd', 'sinhala': 'si', 'slovak': 'sk', 'slovenian': 'sl', 'somali': 'so', 'spanish': 'es', 'sundanese': 'su', 'swahili': 'sw', 'swedish': 'sv', 'tajik': 'tg', 'tamil': 'ta', 'tatar': 'tt', 'telugu': 'te', 'thai': 'th', 'tigrinya': 'ti', 'tsonga': 'ts', 'turkish': 'tr', 'turkmen': 'tk', 'twi': 'ak', 'ukrainian': 'uk', 'urdu': 'ur', 'uyghur': 'ug', 'uzbek': 'uz', 'vietnamese': 'vi', 'welsh': 'cy', 'xhosa': 'xh', 'yiddish': 'yi', 'yoruba': 'yo', 'zulu': 'zu'")
        langue = input("Quelle langue existante voulez choisire dans la liste du haut : ")
        translator = GoogleTranslator(from_lang="fr", to_lang=langue)

        # Créer un nouveau fichier Excel pour écrire les résultats
        new_wb = openpyxl.Workbook()
        new_sheet = new_wb.active

        # Parcourir toutes les cellules de la colonne B
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str):
                    text_to_translate = cell.value.replace(' ', '_') # Remplacer les espaces par des underscores
                    # Traduction du texte en français
                    translation = GoogleTranslator(source="auto", target=langue).translate(text_to_translate)
                    new_sheet.cell(row=row[0].row, column=2).value = translation # Écrire le résultat dans la colonne B du nouveau fichier
                    new_sheet.cell(row=row[0].row, column=1).value = text_to_translate
                    print(translation)
        rowe += 1
        rowee = str(rowe)
    # Enregistrer le nouveau fichier Excel
        new_wb.save(f'New_File{rowe}.xlsx')
    print("Bravo Réussi")


def New_File():
    
    name_of_language = input("Quelle est le nom de votre langue : ")
    namexlsx = name_of_language + ".xlsx"
    print(namexlsx)
    new_file = openpyxl.Workbook()
    aci = new_file.active
    aci['C7'] = "Bonjour"
    new_file.save(namexlsx)

    wb_français = openpyxl.load_workbook('français.xlsx')
    spanglish = openpyxl.load_workbook(namexlsx)

    wb_file = openpyxl.load_workbook("New_File2.xlsx")
    wb_file2 = openpyxl.load_workbook("New_File3.xlsx")

    feuille_français = wb_français['Feuille1']
    f_spanglish = spanglish['Sheet']
    #f_allemend = wb_allemnd['Sheet']
    f_file = wb_file['Sheet']
    f_file2 = wb_file2['Sheet']
    compteur = 1
    for cellule in feuille_français['A']:
        mot_a_chercher = cellule.value
        if mot_a_chercher == None:
            break

        # Parcourir toutes les cellules de la colonne A de 'fichier1.xlsx'
        for row in range(1, feuille_français.max_row + 1):
            cellule_fichier1 = feuille_français.cell(row=row, column=1)
            if cellule_fichier1.value == mot_a_chercher:

                # Chercher le même mot dans la colonne A de 'fichier2.xlsx'
                for row2 in range(1, f_file.max_row + 1):
                    cellule_fichier2 = f_file.cell(row=row2, column=1)
                    if cellule_fichier2.value == mot_a_chercher:
                        français = feuille_français.cell(row=row, column=1).value

                        new_langue = f_file.cell(row=row, column=2).value
                        new_langue2 = f_file2.cell(row=row, column=2).value
                        # Formater le mot
                        new_world = None
                        if new_langue is not None:
                            if len(new_langue) <= 3:
                                new_world = new_langue[0]
                            elif len(new_langue) == 4:
                                new_world = new_langue[0:2]
                            else:
                                new_world = new_langue[:3]

                        if new_langue2 is not None:
                            if len(new_langue2) <= 3:
                                new_word = new_langue2[-1]
                            elif len(new_langue2) == 4:
                                new_word = new_langue2[-2:]
                            else:
                                new_word = new_langue2[-3:]
                        termNW = new_world[:-1]
                        if new_world[-1] == new_word[0]:
                            if new_world[-1] == "i":
                                   new_world = new_world[:-1] + "y"
                            new_world = new_world[:-1]
                    
                        very_new_word = new_world + new_word
                        compteur = 0
                        
                        very_new_word = very_new_word.lower()
                        # Écrire le mot formaté dans le fichier Spanglish.xlsx
                        f_spanglish.cell(row=row, column=2).value = very_new_word
                        f_spanglish.cell(row=row, column=1).value = français
                        # Afficher un message pour confirmer que le mot a été ajouté
                        print(f"Le mot '{very_new_word}' a été ajouté dans le fichier {namexlsx}")
                        break
                else:
                    print(f"Le mot '{mot_a_chercher}' a été trouvé dans fichier1.xlsx à la ligne {row}, mais pas dans fichier2.xlsx.")
                break

        else:
            print(f"Le mot '{mot_a_chercher}' n'a pas été trouvé dans fichier1.xlsx.")
        compteur += 1

    # Créer le dossier "robe" s'il n'existe pas déjà

    # Sauvegarder les modifications apportées au fichier Spanglish.xlsx
    spanglish.save(namexlsx)

translate()


New_File()

