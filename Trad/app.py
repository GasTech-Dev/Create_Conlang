import openpyxl
from flask import Flask, request, render_template
from deep_translator import GoogleTranslator


def traducteur(motAtraduire):
    name_ofFile = "GrecToumoul"#Marquer Le nom de la langue

    name_ofFile = name_ofFile + ".xlsx"
    chemin_fichier =name_ofFile
    wb_langue = openpyxl.load_workbook(chemin_fichier)
    feuille_langue = wb_langue['Sheet']
    motAtraduire = motAtraduire.split()
    mottr = ""
    for mot in motAtraduire:
        
        for row in range(1, feuille_langue.max_row + 1):
            MotLangue = feuille_langue.cell(row=row, column=1).value
            if MotLangue == mot:
                ligne = row

                for rowe in range(1, feuille_langue.max_row + 1):
                    MotTraduit = feuille_langue.cell(row=row, column=2).value
                    
                    mottr += MotTraduit + " "

                    break
    

        print(mottr)
    if mottr == "" or mottr == " ":
        translation_deux = GoogleTranslator(source="auto", target="es").translate(mot)
        translation_un = GoogleTranslator(source="auto", target="it").translate(mot)
        print(translation_un)
        print(translation_deux)
        if len(translation_un) <= 2:
            one = translation_un[:1]
            print(one)
        elif len(translation_un) <= 3:
            one = translation_un[:2]
            print(one)
        elif len(translation_un) <= 4:
            one = translation_un[:3]
            print(one)
        elif len(translation_un) <= 5:
            one = translation_un[:4]
            print(one)
        elif len(translation_deux) <= 6:
            one = translation_un[:5]
            print(one)
        else:
            one = translation_un[:6]
            print(one)

        if len(translation_deux) <= 2:
            two = translation_deux[1:]
            print(two)
        elif len(translation_deux) <= 3:
            two = translation_deux[2:]
            print(two)
        elif len(translation_deux) <= 4:
            two = translation_deux[3:]
            print(two)
        elif len(translation_deux) <= 5:
            two = translation_deux[4:]
            print(two)
        elif len(translation_deux) <= 6:
            two = translation_deux[5:]
            print(two)
        else:
            two = translation_deux[6:]
            print(two)
        mottr = one+two
    return mottr
    wb_langue.close()  # Fermer le fichier Excel


app = Flask(__name__)

@app.route('/traduct', methods=['GET', 'POST'])
def traduct():
    if request.method == 'POST':
        # Traitez le formulaire POST
        data = request.form.get('data')
        trad = traducteur(data)
        return render_template('index.html', message=f"Mot traduit: {trad}")
    else:
        # Affichez simplement la page avec un message par défaut pour GET
        return render_template('index.html', message="Entrez un mot pour le traduire.")
@app.route('/')
def index():
    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)

#traducteur()